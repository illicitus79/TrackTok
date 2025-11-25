"""Reports and analytics API endpoints."""
import csv
import io
from datetime import datetime, timedelta
from typing import Optional

from flask import Response, g, jsonify, make_response, request, stream_with_context
from flask.views import MethodView
from flask_smorest import Blueprint, abort
from loguru import logger
from sqlalchemy import and_, extract, func

from app.core.extensions import db
from app.models.account import Account
from app.models.category import Category
from app.models.expense import Expense
from app.models.project import Project
from app.utils.decorators import roles_required

blp = Blueprint(
    "reports",
    __name__,
    url_prefix="/api/v1/reports",
    description="Reports and analytics",
)


@blp.route("/project/<int:project_id>/summary")
class ProjectSummaryReport(MethodView):
    """Project summary with burn rate analysis."""
    
    @blp.response(200)
    @roles_required(["admin", "manager", "user"])
    def get(self, project_id):
        """Get project summary report with burn rate calculations."""
        tenant_id = g.current_tenant_id
        
        # Get project
        project = Project.query.filter_by(
            id=project_id, tenant_id=tenant_id
        ).first()
        
        if not project:
            abort(404, message="Project not found")
        
        # Get all expenses for the project
        expenses = Expense.query.filter_by(
            project_id=project_id,
            tenant_id=tenant_id
        ).all()
        
        # Calculate totals
        total_spend = sum(e.amount for e in expenses)
        starting_budget = project.starting_budget or 0
        projected_estimate = project.projected_estimate or starting_budget
        remaining_budget = starting_budget - total_spend
        expense_count = len(expenses)
        budget_utilization = (total_spend / starting_budget * 100) if starting_budget > 0 else 0
        
        # Calculate burn rate (last 30 days)
        now = datetime.utcnow()
        start_date = now - timedelta(days=30)
        
        recent_expenses = [e for e in expenses if e.expense_date >= start_date]
        days_elapsed = (now - start_date).days or 1
        
        daily_burn_rate = total_spend / days_elapsed if days_elapsed > 0 else 0
        monthly_burn_rate = daily_burn_rate * 30
        
        # Calculate projected completion
        projected_completion_date = None
        days_remaining = None
        
        if daily_burn_rate > 0 and remaining_budget > 0:
            days_remaining = int(remaining_budget / daily_burn_rate)
            projected_completion_date = (now + timedelta(days=days_remaining)).isoformat()
        
        # Build burn rate chart data (daily spend over last 30 days)
        chart_data = []
        cumulative = 0
        
        for i in range(30):
            date = start_date + timedelta(days=i)
            date_str = date.strftime("%Y-%m-%d")
            
            daily_expenses = [e for e in recent_expenses if e.expense_date.date() == date.date()]
            daily_total = sum(e.amount for e in daily_expenses)
            cumulative += daily_total
            
            chart_data.append({
                "date": date_str,
                "daily": round(daily_total, 2),
                "cumulative": round(cumulative, 2)
            })
        
        return {
            "project": {
                "id": project.id,
                "name": project.name,
                "status": project.status,
                "start_date": project.start_date.isoformat() if project.start_date else None,
                "end_date": project.end_date.isoformat() if project.end_date else None
            },
            "totals": {
                "starting_budget": round(starting_budget, 2),
                "projected_estimate": round(projected_estimate, 2),
                "total_spend": round(total_spend, 2),
                "remaining_budget": round(remaining_budget, 2),
                "expense_count": expense_count,
                "budget_utilization": round(budget_utilization, 2)
            },
            "burn_rate": {
                "daily": round(daily_burn_rate, 2),
                "monthly": round(monthly_burn_rate, 2),
                "projected_completion_date": projected_completion_date,
                "days_remaining": days_remaining
            },
            "burn_rate_chart": {
                "labels": [d["date"] for d in chart_data],
                "daily": [d["daily"] for d in chart_data],
                "cumulative": [d["cumulative"] for d in chart_data]
            }
        }


@blp.route("/project/<int:project_id>/category-breakdown")
class ProjectCategoryBreakdownReport(MethodView):
    """Category breakdown for a project."""
    
    @blp.response(200)
    @roles_required(["admin", "manager", "user"])
    def get(self, project_id):
        """Get category breakdown with donut chart data."""
        tenant_id = g.current_tenant_id
        
        # Verify project exists
        project = Project.query.filter_by(
            id=project_id, tenant_id=tenant_id
        ).first()
        
        if not project:
            abort(404, message="Project not found")
        
        # Get category breakdown
        breakdown_query = (
            db.session.query(
                Category.id.label("category_id"),
                Category.name.label("category_name"),
                Category.color.label("category_color"),
                func.sum(Expense.amount).label("total"),
                func.count(Expense.id).label("count")
            )
            .select_from(Expense)
            .outerjoin(Category, Expense.category_id == Category.id)
            .filter(
                Expense.project_id == project_id,
                Expense.tenant_id == tenant_id
            )
            .group_by(Category.id, Category.name, Category.color)
        )
        
        results = breakdown_query.all()
        
        # Calculate total spend
        total_spend = sum(r.total for r in results)
        
        # Build breakdown array
        breakdown = []
        for r in results:
            percentage = (r.total / total_spend * 100) if total_spend > 0 else 0
            breakdown.append({
                "category_id": r.category_id,
                "category_name": r.category_name or "Uncategorized",
                "color": r.category_color or "#6b7280",
                "total": round(r.total, 2),
                "count": r.count,
                "percentage": round(percentage, 2)
            })
        
        # Sort by total descending
        breakdown.sort(key=lambda x: x["total"], reverse=True)
        
        # Build chart data
        chart = {
            "labels": [b["category_name"] for b in breakdown],
            "data": [b["total"] for b in breakdown],
            "colors": [b["color"] for b in breakdown],
            "percentages": [b["percentage"] for b in breakdown]
        }
        
        return {
            "project_id": project_id,
            "total_spend": round(total_spend, 2),
            "breakdown": breakdown,
            "chart": chart
        }


@blp.route("/project/<int:project_id>/monthly-trend")
class ProjectMonthlyTrendReport(MethodView):
    """Monthly trend report with stacked datasets."""
    
    @blp.response(200)
    @roles_required(["admin", "manager", "user"])
    def get(self, project_id):
        """Get monthly trend report grouped by category or account."""
        tenant_id = g.current_tenant_id
        
        # Query params
        year = request.args.get("year", datetime.utcnow().year, type=int)
        group_by = request.args.get("group_by", "category")  # category or account
        
        # Verify project exists
        project = Project.query.filter_by(
            id=project_id, tenant_id=tenant_id
        ).first()
        
        if not project:
            abort(404, message="Project not found")
        
        # Build query based on group_by
        if group_by == "category":
            trend_query = (
                db.session.query(
                    extract("month", Expense.expense_date).label("month"),
                    Category.id.label("group_id"),
                    Category.name.label("group_name"),
                    Category.color.label("group_color"),
                    func.sum(Expense.amount).label("total")
                )
                .select_from(Expense)
                .outerjoin(Category, Expense.category_id == Category.id)
                .filter(
                    Expense.project_id == project_id,
                    Expense.tenant_id == tenant_id,
                    extract("year", Expense.expense_date) == year
                )
                .group_by("month", Category.id, Category.name, Category.color)
                .order_by("month")
            )
        else:  # account
            trend_query = (
                db.session.query(
                    extract("month", Expense.expense_date).label("month"),
                    Account.id.label("group_id"),
                    Account.name.label("group_name"),
                    func.sum(Expense.amount).label("total")
                )
                .select_from(Expense)
                .join(Account, Expense.account_id == Account.id)
                .filter(
                    Expense.project_id == project_id,
                    Expense.tenant_id == tenant_id,
                    extract("year", Expense.expense_date) == year
                )
                .group_by("month", Account.id, Account.name)
                .order_by("month")
            )
        
        results = trend_query.all()
        
        # Initialize 12 months
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        monthly_totals = [0] * 12
        
        # Group data by group_id
        groups = {}
        for r in results:
            month_idx = int(r.month) - 1
            group_id = r.group_id or 0
            group_name = r.group_name or "Uncategorized"
            
            if group_id not in groups:
                groups[group_id] = {
                    "name": group_name,
                    "color": getattr(r, "group_color", None) or "#6b7280",
                    "data": [0] * 12
                }
            
            groups[group_id]["data"][month_idx] += r.total
            monthly_totals[month_idx] += r.total
        
        # Build datasets for stacked bar chart
        datasets = []
        for group_id, group_data in groups.items():
            datasets.append({
                "label": group_data["name"],
                "data": [round(v, 2) for v in group_data["data"]],
                "backgroundColor": group_data["color"]
            })
        
        # Calculate summary
        total_spend = sum(monthly_totals)
        average_monthly = total_spend / 12
        highest_month = max(monthly_totals)
        lowest_month = min([m for m in monthly_totals if m > 0], default=0)
        
        return {
            "project_id": project_id,
            "year": year,
            "group_by": group_by,
            "chart": {
                "labels": months,
                "datasets": datasets,
                "monthly_totals": [round(v, 2) for v in monthly_totals]
            },
            "summary": {
                "total_spend": round(total_spend, 2),
                "average_monthly": round(average_monthly, 2),
                "highest_month": round(highest_month, 2),
                "lowest_month": round(lowest_month, 2)
            }
        }


@blp.route("/tenant/cashflow")
class TenantCashflowReport(MethodView):
    """Tenant-wide cashflow report."""
    
    @blp.response(200)
    @roles_required(["admin", "manager"])
    def get(self):
        """Get tenant cashflow with inflow/outflow analysis."""
        tenant_id = g.current_tenant_id
        
        # Query params for date range
        from_date_str = request.args.get("from")
        to_date_str = request.args.get("to")
        
        # Default to last 90 days
        to_date = datetime.utcnow()
        from_date = to_date - timedelta(days=90)
        
        if from_date_str:
            try:
                from_date = datetime.fromisoformat(from_date_str)
            except ValueError:
                abort(400, message="Invalid from date format")
        
        if to_date_str:
            try:
                to_date = datetime.fromisoformat(to_date_str)
            except ValueError:
                abort(400, message="Invalid to date format")
        
        # Get daily outflow (expenses)
        daily_query = (
            db.session.query(
                func.date(Expense.expense_date).label("date"),
                func.sum(Expense.amount).label("outflow")
            )
            .filter(
                Expense.tenant_id == tenant_id,
                Expense.expense_date >= from_date,
                Expense.expense_date <= to_date
            )
            .group_by(func.date(Expense.expense_date))
            .order_by("date")
        )
        
        results = daily_query.all()
        
        # Build daily data
        daily_data = {}
        for r in results:
            daily_data[r.date.isoformat()] = {
                "inflow": 0,  # Future feature
                "outflow": r.outflow
            }
        
        # Fill in missing dates
        current_date = from_date.date()
        end_date = to_date.date()
        
        chart_labels = []
        inflow_data = []
        outflow_data = []
        net_data = []
        cumulative_data = []
        cumulative = 0
        
        while current_date <= end_date:
            date_str = current_date.isoformat()
            chart_labels.append(date_str)
            
            day_data = daily_data.get(date_str, {"inflow": 0, "outflow": 0})
            inflow = day_data["inflow"]
            outflow = day_data["outflow"]
            net = inflow - outflow
            cumulative += net
            
            inflow_data.append(round(inflow, 2))
            outflow_data.append(round(outflow, 2))
            net_data.append(round(net, 2))
            cumulative_data.append(round(cumulative, 2))
            
            current_date += timedelta(days=1)
        
        # Calculate summary
        total_inflow = sum(inflow_data)
        total_outflow = sum(outflow_data)
        net_cashflow = total_inflow - total_outflow
        days = len(chart_labels)
        average_daily_outflow = total_outflow / days if days > 0 else 0
        
        return {
            "period": {
                "from": from_date.isoformat(),
                "to": to_date.isoformat()
            },
            "summary": {
                "total_inflow": round(total_inflow, 2),
                "total_outflow": round(total_outflow, 2),
                "net_cashflow": round(net_cashflow, 2),
                "days": days,
                "average_daily_outflow": round(average_daily_outflow, 2)
            },
            "chart": {
                "labels": chart_labels,
                "inflow": inflow_data,
                "outflow": outflow_data,
                "net": net_data,
                "cumulative": cumulative_data
            }
        }


@blp.route("/export/csv")
@roles_required(["admin", "manager", "user"])
def export_csv():
    """Export expenses to CSV."""
    tenant_id = g.current_tenant_id
    
    # Query params for filtering
    project_id = request.args.get("project_id", type=int)
    from_date_str = request.args.get("from")
    to_date_str = request.args.get("to")
    
    # Build query
    query = Expense.query.filter_by(tenant_id=tenant_id)
    
    if project_id:
        query = query.filter_by(project_id=project_id)
    
    if from_date_str:
        try:
            from_date = datetime.fromisoformat(from_date_str)
            query = query.filter(Expense.expense_date >= from_date)
        except ValueError:
            pass
    
    if to_date_str:
        try:
            to_date = datetime.fromisoformat(to_date_str)
            query = query.filter(Expense.expense_date <= to_date)
        except ValueError:
            pass
    
    expenses = query.order_by(Expense.expense_date.desc()).all()
    
    # Generate CSV
    def generate():
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Date", "Vendor", "Amount", "Currency", "Category", 
            "Project", "Account", "Note", "Created At"
        ])
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
        
        # Data rows
        for expense in expenses:
            writer.writerow([
                expense.expense_date.strftime("%Y-%m-%d"),
                expense.vendor or "",
                expense.amount,
                expense.currency or "USD",
                expense.category.name if expense.category else "",
                expense.project.name if expense.project else "",
                expense.account.name if expense.account else "",
                expense.note or "",
                expense.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)
    
    response = Response(stream_with_context(generate()), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=expenses_export.csv"
    
    return response


@blp.route("/export/xlsx")
@roles_required(["admin", "manager", "user"])
def export_xlsx():
    """Export expenses to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        abort(501, message="openpyxl not installed. Run: pip install openpyxl")
    
    tenant_id = g.current_tenant_id
    
    # Query params for filtering
    project_id = request.args.get("project_id", type=int)
    from_date_str = request.args.get("from")
    to_date_str = request.args.get("to")
    
    # Build query
    query = Expense.query.filter_by(tenant_id=tenant_id)
    
    if project_id:
        query = query.filter_by(project_id=project_id)
    
    if from_date_str:
        try:
            from_date = datetime.fromisoformat(from_date_str)
            query = query.filter(Expense.expense_date >= from_date)
        except ValueError:
            pass
    
    if to_date_str:
        try:
            to_date = datetime.fromisoformat(to_date_str)
            query = query.filter(Expense.expense_date <= to_date)
        except ValueError:
            pass
    
    expenses = query.order_by(Expense.expense_date.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Header styling
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["Date", "Vendor", "Amount", "Currency", "Category", 
               "Project", "Account", "Note", "Created At"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row_idx, expense in enumerate(expenses, 2):
        ws.cell(row=row_idx, column=1, value=expense.expense_date.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=2, value=expense.vendor or "")
        ws.cell(row=row_idx, column=3, value=expense.amount)
        ws.cell(row=row_idx, column=4, value=expense.currency or "USD")
        ws.cell(row=row_idx, column=5, value=expense.category.name if expense.category else "")
        ws.cell(row=row_idx, column=6, value=expense.project.name if expense.project else "")
        ws.cell(row=row_idx, column=7, value=expense.account.name if expense.account else "")
        ws.cell(row=row_idx, column=8, value=expense.note or "")
        ws.cell(row=row_idx, column=9, value=expense.created_at.strftime("%Y-%m-%d %H:%M:%S"))
    
    # Adjust column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = "attachment; filename=expenses_export.xlsx"
    
    return response
